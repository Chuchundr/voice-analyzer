from openpyxl import load_workbook
from openpyxl.styles import Font

WORKBOOK_NAME = 'FLM Uzmobile Список БС 1000 new.xlsx'

font = Font(name='Times New Roman')


class ExcelClass:
    def __init__(self):
        self.wb = load_workbook(WORKBOOK_NAME)
        self.sheet = self.wb.get_sheet_by_name('Site visit Tracker')

    def write_into_cell(self, column, row, text):
        mycell = self.sheet.cell(row, column)
        mycell.value = text
        self.wb.save(WORKBOOK_NAME)